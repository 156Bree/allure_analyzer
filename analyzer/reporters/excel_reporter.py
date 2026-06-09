"""excel_reporter：导出单个多 sheet 工作簿（需 openpyxl）。

缺少 openpyxl 时优雅跳过（返回 None 并打印提示），不影响其他格式。
"""
import os


def _pct(x):
    return "%.2f%%" % (x * 100)


def write(analysis, reconcile_result, out_dir):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[excel] 未安装 openpyxl，跳过 xlsx 导出（pip install openpyxl 可启用）。")
        return None

    wb = Workbook()
    wb.remove(wb.active)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")

    def add_sheet(title, header, rows, status_col=None):
        ws = wb.create_sheet(title[:31])
        ws.append(header)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
            if status_col is not None:
                cell = ws.cell(row=ws.max_row, column=status_col + 1)
                if str(cell.value) in ("Pass", "PASS"):
                    cell.fill = pass_fill
                elif str(cell.value) in ("Fail", "FAIL"):
                    cell.fill = fail_fill
        # 自适应列宽（粗略）
        for i, _h in enumerate(header, 1):
            maxlen = max([len(str(header[i - 1]))] +
                         [len(str(r[i - 1])) for r in rows if i - 1 < len(r)] or [0])
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(maxlen + 2, 60)
        ws.freeze_panes = "A2"
        return ws

    ov = analysis["overview"]
    add_sheet("Overview", ["grain", "total", "pass", "fail", "pass_rate"], [
        ["scenario", ov["scenario"]["total"], ov["scenario"]["pass"],
         ov["scenario"]["fail"], _pct(ov["scenario"]["pass_rate"])],
        ["case", ov["case"]["total"], ov["case"]["pass"],
         ov["case"]["fail"], _pct(ov["case"]["pass_rate"])],
    ])

    add_sheet("Case Result",
              ["case_key", "jira_url", "owners", "components", "host_display",
               "verdict", "n_pass", "n_total"],
              [[r["case_key"], r["jira_url"], "|".join(r["owners"]),
                "|".join(r["components"]), r["host_display"], r["verdict"],
                r["n_pass"], r["n_total"]] for r in analysis["case_host_rows"]],
              status_col=5)

    for grain, key in (("Case", "case_summary"), ("Step", "scenario_summary")):
        for dim in ("owner", "component", "device"):
            add_sheet("%s Sum-%s" % (grain, dim),
                      [dim, "pass", "fail", "total", "pass_rate"],
                      [[r["name"], r["pass"], r["fail"], r["total"], _pct(r["pass_rate"])]
                       for r in analysis[key][dim]])

    mx = analysis["matrix"]
    header = ["host\\owner"] + mx["owners"]
    rows = []
    for h in mx["hosts"]:
        row = [h]
        for o in mx["owners"]:
            c = mx["cells"].get("%s||%s" % (h, o))
            row.append("%d/%d" % (c["pass"], c["total"]) if c else "")
        rows.append(row)
    add_sheet("Matrix host x owner", header, rows)

    add_sheet("Failures",
              ["case_key", "host", "scenario", "status", "severity", "status_message"],
              [[r["case_key"], r["host"], r["scenario"], r["status"], r["severity"],
                (r["status_message"] or "").replace("\n", " ")[:500]]
               for r in analysis["failures"]])

    add_sheet("Inconsistent", ["case_key", "jira_url", "host_verdicts"],
              [[r["case_key"], r["jira_url"],
                "; ".join("%s=%s" % (h, v) for h, v in r["hosts"].items())]
               for r in analysis["inconsistent"]])

    add_sheet("Top Duration", ["case_key", "host", "scenario", "duration_ms"],
              [[r["case_key"], r["host"], r["scenario"], r["duration_ms"]]
               for r in analysis["top_duration"]])
    add_sheet("Top Retries", ["case_key", "host", "scenario", "retries_count"],
              [[r["case_key"], r["host"], r["scenario"], r["retries_count"]]
               for r in analysis["top_retries"]])

    add_sheet("Key Source", ["key_source", "count"],
              [[k, v] for k, v in analysis["key_source_dist"].items()])

    rr = reconcile_result
    add_sheet("Reconcile", ["metric", "value"], [
        ["records_total", rr["records_total"]],
        ["visible_records", rr["visible_records"]],
        ["hidden_records", rr["hidden_records"]],
        ["summary_json_total", rr["summary_total"]],
        ["match", rr["match"]],
        ["note", rr["note"]],
    ])

    path = os.path.join(out_dir, "analysis.xlsx")
    wb.save(path)
    print("[excel] 已写出 %s" % path)
    return path
